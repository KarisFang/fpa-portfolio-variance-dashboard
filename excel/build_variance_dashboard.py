"""
Excel Budget vs. Actual Variance Dashboard.
Variance % is computed with live Excel formulas; conditional formatting
(color scales + rule-based highlighting) does the "flag it visually" work.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

FONT = "Calibri"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
BLACK_BOLD = Font(name=FONT, color="000000", bold=True)
WHITE_BOLD = Font(name=FONT, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT, size=10, color="595959")
KPI_VALUE_FONT = Font(name=FONT, bold=True, size=18, color="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
KPI_FILL = PatternFill("solid", fgColor="F2F2F2")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(name=FONT, color="9C0006")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(name=FONT, color="006100")
CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

EXPORT_DIR = "/home/claude/fpa-variance-dashboard/exports"

wb = Workbook()
wb.remove(wb.active)


def load_csv(name):
    with open(f"{EXPORT_DIR}/{name}") as f:
        r = list(csv.reader(f))
    return r[0], r[1:]


# =====================================================================
# SHEET: Data_Region_Variance — region x month, budget/actual hardcoded,
# variance $ and % computed live in Excel, conditional formatting on %.
# =====================================================================
ws = wb.create_sheet("Region_Variance")
ws["A1"] = "Revenue Variance by Region x Month — FY2025"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Budget = FY2024 actual x planned +10% growth, set before the year started. Source: SQL export."
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")

headers = ["Region", "Month", "Budget", "Actual", "Variance $", "Variance %", "Exception"]
for j, h in enumerate(headers):
    c = ws.cell(row=4, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL

hdr, rows = load_csv("region_variance_monthly.csv")
for i, row in enumerate(rows):
    r = 5 + i
    region, month, budget, actual = row
    ws.cell(row=r, column=1, value=region).font = BLUE
    ws.cell(row=r, column=2, value=int(month)).font = BLUE
    ws.cell(row=r, column=3, value=float(budget)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=float(actual)).font = BLUE
    ws.cell(row=r, column=4).number_format = CUR
    ws.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = CUR
    ws.cell(row=r, column=5).font = BLACK
    ws.cell(row=r, column=6, value=f"=D{r}/C{r}-1").number_format = PCT
    ws.cell(row=r, column=6).font = BLACK
    ws.cell(row=r, column=7, value=f'=IF(ABS(F{r})>0.1,"FLAG","")').font = BLACK

last_row = 4 + len(rows)

# Conditional formatting: 3-color scale on Variance % (red=bad/low, green=good/high)
ws.conditional_formatting.add(
    f"F5:F{last_row}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B")
)
# Bold-red highlight any |variance| > 10% exception
ws.conditional_formatting.add(
    f"F5:F{last_row}",
    FormulaRule(formula=[f"ABS(F5)>0.1"], font=RED_FONT, fill=RED_FILL)
)
ws.conditional_formatting.add(
    f"G5:G{last_row}",
    CellIsRule(operator="equal", formula=['"FLAG"'], font=RED_FONT, fill=RED_FILL)
)

for col, w in {"A": 16, "B": 8, "C": 14, "D": 14, "E": 14, "F": 12, "G": 11}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"
print("Region_Variance done:", last_row)

# =====================================================================
# SHEET: Department_Variance — OpEx, department x month
# Sign convention note: for costs, positive variance % = overspend (bad)
# so the color scale is REVERSED relative to revenue.
# =====================================================================
ws = wb.create_sheet("Department_Variance")
ws["A1"] = "OpEx Variance by Department x Month — FY2025"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Positive variance % = overspend vs. budget (unfavorable). Source: SQL export."
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")

headers = ["Department", "Month", "Budget", "Actual", "Variance $", "Variance %", "Exception"]
for j, h in enumerate(headers):
    c = ws.cell(row=4, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL

hdr, rows = load_csv("department_variance_monthly.csv")
for i, row in enumerate(rows):
    r = 5 + i
    dept, month, budget, actual = row
    ws.cell(row=r, column=1, value=dept).font = BLUE
    ws.cell(row=r, column=2, value=int(month)).font = BLUE
    ws.cell(row=r, column=3, value=float(budget)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=float(actual)).font = BLUE
    ws.cell(row=r, column=4).number_format = CUR
    ws.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = CUR
    ws.cell(row=r, column=5).font = BLACK
    ws.cell(row=r, column=6, value=f"=D{r}/C{r}-1").number_format = PCT
    ws.cell(row=r, column=6).font = BLACK
    ws.cell(row=r, column=7, value=f'=IF(F{r}>0.05,"OVERSPEND",IF(F{r}<-0.05,"UNDERSPEND",""))').font = BLACK

last_row_d = 4 + len(rows)

# Reversed color scale: green = underspend/favorable (low %), red = overspend (high %)
ws.conditional_formatting.add(
    f"F5:F{last_row_d}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   mid_type="num", mid_value=0, mid_color="FFEB84",
                   end_type="max", end_color="F8696B")
)
ws.conditional_formatting.add(
    f"F5:F{last_row_d}",
    FormulaRule(formula=[f"F5>0.05"], font=RED_FONT, fill=RED_FILL)
)
ws.conditional_formatting.add(
    f"G5:G{last_row_d}",
    CellIsRule(operator="equal", formula=['"OVERSPEND"'], font=RED_FONT, fill=RED_FILL)
)
ws.conditional_formatting.add(
    f"G5:G{last_row_d}",
    CellIsRule(operator="equal", formula=['"UNDERSPEND"'], font=GREEN_FONT, fill=GREEN_FILL)
)

for col, w in {"A": 22, "B": 8, "C": 13, "D": 13, "E": 13, "F": 12, "G": 12}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"
print("Department_Variance done:", last_row_d)

# =====================================================================
# SHEET: Dashboard
# =====================================================================
ws = wb.create_sheet("Dashboard")
ws["A1"] = "Meridian Retail Co. — Budget vs. Actual Variance Dashboard (FY2025)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:N1")
ws["A2"] = "Drill down on Region_Variance and Department_Variance tabs (conditional formatting flags every exception)"
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A2:N2")

kpis = [
    ("FY2025 Revenue Variance $", f"=SUMIF(Region_Variance!A5:A{last_row},\"<>\")*0+SUM(Region_Variance!E5:E{last_row})", CUR),
    ("FY2025 Revenue Variance %", f"=SUM(Region_Variance!D5:D{last_row})/SUM(Region_Variance!C5:C{last_row})-1", PCT),
    ("FY2025 OpEx Variance $ (Favorable)", f"=-SUM(Department_Variance!E5:E{last_row_d})", CUR),
    ("FY2025 OpEx Variance %", f"=SUM(Department_Variance!D5:D{last_row_d})/SUM(Department_Variance!C5:C{last_row_d})-1", PCT),
    ("Flagged Exceptions (Rev + OpEx)", f"=COUNTIF(Region_Variance!G5:G{last_row},\"FLAG\")+COUNTIF(Department_Variance!G5:G{last_row_d},\"OVERSPEND\")+COUNTIF(Department_Variance!G5:G{last_row_d},\"UNDERSPEND\")", "0"),
]
for i, (lbl, formula, fmt) in enumerate(kpis):
    col = 1 + i * 3
    cell_lbl = ws.cell(row=4, column=col, value=lbl)
    cell_lbl.font = KPI_LABEL_FONT
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    cell_val = ws.cell(row=5, column=col, value=formula)
    cell_val.font = KPI_VALUE_FONT
    cell_val.number_format = fmt
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    for rr in (4, 5):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).fill = KPI_FILL
    ws.row_dimensions[5].height = 26

# --- Annual region summary table (via SUMIF over Region_Variance) ---
ws["A7"] = "Revenue: Budget vs. Actual by Region (FY2025 Full Year)"
ws["A7"].font = BLACK_BOLD
reg_headers = ["Region", "Budget", "Actual", "Variance %"]
for j, h in enumerate(reg_headers):
    c = ws.cell(row=8, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
hdr, reg_rows = load_csv("region_variance_annual.csv")
for i, row in enumerate(reg_rows):
    r = 9 + i
    region = row[0]
    ws.cell(row=r, column=1, value=region).font = BLACK
    ws.cell(row=r, column=2, value=f'=SUMIF(Region_Variance!$A$5:$A${last_row},A{r},Region_Variance!$C$5:$C${last_row})').number_format = CUR
    ws.cell(row=r, column=3, value=f'=SUMIF(Region_Variance!$A$5:$A${last_row},A{r},Region_Variance!$D$5:$D${last_row})').number_format = CUR
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = PCT
reg_last = 8 + len(reg_rows)
ws.conditional_formatting.add(
    f"D9:D{reg_last}",
    ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
                   mid_color="FFEB84", end_type="max", end_color="63BE7B")
)

# --- Annual department summary table ---
ws["F7"] = "OpEx: Budget vs. Actual by Department (FY2025 Full Year)"
ws["F7"].font = BLACK_BOLD
dep_headers = ["Department", "Budget", "Actual", "Variance %"]
for j, h in enumerate(dep_headers):
    c = ws.cell(row=8, column=6 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
hdr, dep_rows = load_csv("department_variance_annual.csv")
for i, row in enumerate(dep_rows):
    r = 9 + i
    dept = row[0]
    ws.cell(row=r, column=6, value=dept).font = BLACK
    ws.cell(row=r, column=7, value=f'=SUMIF(Department_Variance!$A$5:$A${last_row_d},F{r},Department_Variance!$C$5:$C${last_row_d})').number_format = CUR
    ws.cell(row=r, column=8, value=f'=SUMIF(Department_Variance!$A$5:$A${last_row_d},F{r},Department_Variance!$D$5:$D${last_row_d})').number_format = CUR
    ws.cell(row=r, column=9, value=f"=H{r}/G{r}-1").number_format = PCT
dep_last = 8 + len(dep_rows)
ws.conditional_formatting.add(
    f"I9:I{dep_last}",
    ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="num", mid_value=0,
                   mid_color="FFEB84", end_type="max", end_color="F8696B")
)

# --- Charts ---
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Revenue: Budget vs. Actual by Region"
chart1.style = 10
chart1.height = 8.5
chart1.width = 16
data = Reference(ws, min_col=2, max_col=3, min_row=8, max_row=reg_last)
cats = Reference(ws, min_col=1, min_row=9, max_row=reg_last)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws.add_chart(chart1, "A17")

chart2 = BarChart()
chart2.type = "col"
chart2.title = "OpEx: Budget vs. Actual by Department"
chart2.style = 10
chart2.height = 8.5
chart2.width = 16
data = Reference(ws, min_col=7, max_col=8, min_row=8, max_row=dep_last)
cats = Reference(ws, min_col=6, min_row=9, max_row=dep_last)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws.add_chart(chart2, "I17")

ws["A35"] = ("Headline insight: the FY2025 plan assumed a uniform +10% revenue lift across every region. "
             "Greater Asia missed that plan by -43.5% while North America beat it by +12.6% \u2014 a single "
             "region's miss erased most of the company's outperformance elsewhere. On the cost side, G&A "
             "overspend crept up every single month without ever triggering a single-month alarm \u2014 exactly "
             "the kind of slow-creep variance a monthly review (instead of only an annual one) is built to catch.")
ws["A35"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A35:N35")
ws.row_dimensions[35].height = 45

ws.sheet_view.showGridLines = False
for col in "ABCDEFGHIJKLMN":
    ws.column_dimensions[col].width = 11

order = ["Dashboard", "Region_Variance", "Department_Variance"]
wb._sheets = [wb[name] for name in order]
wb.active = 0
print("Dashboard done")
wb.save("/home/claude/fpa-variance-dashboard/excel/Variance_Dashboard.xlsx")
